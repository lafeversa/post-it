import openpyxl as pxl
import logging as log
from os.path import exists
from os import replace
from active_check import run as active_run

def compare_checks():
    '''Compares data from active_file to data in master_file and stores new data as a list.
    '''
    ids = load_ids()
    mids = load_mids()
    
    nids = []
    
    for id in ids:
        if id not in mids:
            nids.append(id)
    
    print(nids)
    
    return nids
    
def load_mids(master_file='master_file.xlsx'):
    '''Loads all the unique identifiers for posted jobs found in master_file.
    '''
    m_WB = pxl.load_workbook(master_file)
    m_ST = m_WB.active
    num_rows = len(m_ST.rows)
    row_count = 2
    col_count = 2
    mids = []
    
    while row_count <= num_rows:
        c_index = m_ST.cell(row = row_count, column = col_count)
        c_val = c_index.value
        if c_val != None:
            mids.append(c_val)
        row_count += 1
        
    return mids

def load_ids(active_file='active_file.xlsx'):
    '''Loads all the unique identifiers for active jobs found in active_file.
    '''
    a_WB = pxl.load_workbook(active_file)
    a_ST = a_WB.active
    num_rows = len(a_ST.rows)
    row_count = 2
    col_count = 2
    ids = []
    
    while row_count <= num_rows:
        c_index = a_ST.cell(row = row_count, column = col_count)
        c_val = c_index.value
        if c_val != None:
            ids.append(c_val)
        row_count += 1

    return ids
    
def find_active_file(active_file='active_file.xlsx'):
    '''Runs the active_check.py script to generate the active_file and proceeds.
    
    Args:
        active_file (str) = File name of active data from user feed. Should end with '.xlsx'. Defaults to 'active_file.xlsx'.
    '''
    active_run()
    load_ids()

def write_blank(fout='master_file.xlsx'):
    '''Writes blank file and saves it to 'fout'.
    
    Args:
        fout (str): File name to save the blank Excel master_file as. Should end with '.xlsx'. Defaults to 'master_file.xlsx'.
    '''
    newWB = pxl.Workbook()
    sheet = newWB.active

    headers = ['Title', 'Link', 'Description', 'Pub Date']
    
    for h in range(len(headers)):
        sheet.cell(row = 1, column = h + 1).value = headers[h]
        
    newWB.save(fout)

def master_check(master_file='master_file.xlsx'):
    '''Checks that master_file exists and proceeds or creates master_file and proceeds.
    
    Args:
        master_file (str) = File name of master data repository. Should end with '.xlsx'. Defaults to 'master_file.xlsx'.
    '''
    if exists(master_file):
        find_active_file()
    else:
        write_blank()
        find_active_file()
        
def run():
    master_check()
    compare_checks()
    
if __name__ == '__main__':
    run()