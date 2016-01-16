from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl as pxl
import random
import time
from collections import OrderedDict
from os.path import exists
from os import replace
import pickle
import logging as log
import urllib.request
import xml.etree.ElementTree as ET

def pull_feed(feed_url='http://cepamerica.force.com/careers/ts2__apply_to_job?nostate=1&tSource=a0dG0000002vObVIAU&showJobs=500',
            filename='active_check.xml'):
    '''
    Retrieves data from a URL and stores it as a file.
    
    Args:
        feed_url (str): Location that contains the data the user is pulling. Defaults to CEP salesforce job link.
        filename (str): File name where the pulled data will be dumped. Defaults to 'active_check.xml'.
    
    Returns:
        The file type and name per user specification.
    '''
    feed_data_pull = urllib.request.urlretrieve(feed_url, filename)
    return filename
    
def parse_xml(filename='active_check.xml'):
    feed = ET.parse(filename)
    root = feed.getroot()
    
    elements = []
    
    for root_elements in root:
        for branch_elements in root_elements:
            post_elements = []
            for leaf_elements in branch_elements:
                post_elements.append(leaf_elements.text)
            elements.append(post_elements)
    return elements
        
def clean_parse_data(filename='active_check.xml'):
    '''Parses the data in `filename` and returns a list of lists containing 
    the data elements for each non-empty post element.
    '''
    elements = parse_xml(filename)
    for item in elements:
        if item == []:
            elements.remove(item)
    return elements

def write_master(fout='master_file.xlsx', filename='active_check.xml'):
    '''Retrieves data from `filename` and saves it to `fout`.
    
    Args:
        fout (str): File name to save the Excel data as. Should end with '.xlsx'. Defaults to 'master_file.xlsx'.
        filename (str): File name containing the xml data. Defaults to 'active_check.xml'.
    '''
    master_data = clean_parse_data(filename)
    newWB = pxl.Workbook()
    sheet = newWB.active
    
    headers = ['Title', 'Link', 'Description', 'Pub Date']
    
    for h in range(len(headers)):
        sheet.cell(row = 1, column = h + 1).value = headers[h]
    
    row_index = 2
    for job in master_data:
        for j in job:
            col_index = job.index(j) + 1
            sheet.cell(row = row_index, column = col_index).value = j
        row_index += 1

    newWB.save(fout)
    
    
def run():
    feed_file = 'active_check.xml'
    save_file = 'master_file.xlsx'
    
    pull_feed(filename=feed_file)
    write_master(fout=save_file, filename=feed_file)

if __name__ == '__main__':
    run()
